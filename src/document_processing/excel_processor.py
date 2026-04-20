# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import os
from io import StringIO
from typing import Any, Dict, List

from .processor_utils import stable_doc_id


class ExcelProcessor:
    """XLSX/XLSM ingestion via openpyxl; CSV/TSV supported directly."""

    def process(self, abs_path: str, rel_path: str, export_root: str, *, doc_id_override: str | None = None) -> Dict[str, Any]:
        doc_id = doc_id_override or stable_doc_id(abs_path, prefix="xlsx")
        ext = os.path.splitext(abs_path)[1].lower()

        text = ""
        structured: List[Dict[str, Any]] = []

        # caps to prevent exploding DB/content_text
        max_rows = None
        max_cols = None

        if ext in {".csv", ".tsv"}:
            delimiter = "," if ext == ".csv" else "\t"
            try:
                with open(abs_path, "r", encoding="utf-8", errors="replace") as f:
                    for i, line in enumerate(f):
                        if max_rows and i >= max_rows:
                            break
                        text += line
            except Exception as e:
                text = f"[tabular] failed to read {ext}: {e}"
        else:
            try:
                import openpyxl

                wb = openpyxl.load_workbook(abs_path, data_only=True) # Removed read_only to allow image extraction
                buf = StringIO()
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    buf.write(f"## Sheet: {sname}\n")
                    
                    try:
                        for idx, img in enumerate(getattr(ws, '_images', [])):
                            img_path = os.path.join(export_root, f"excel_{sname}_img_{idx}.png")
                            with open(img_path, "wb") as f:
                                f.write(img._data())
                            from src.vision.ocr_backends import TesseractBackend
                            ocr_text = TesseractBackend().recognize(img_path, lang_hint="eng")
                            if ocr_text:
                                buf.write(f"[Image OCR] {ocr_text.strip()}\n")
                    except Exception as e:
                        pass

                    for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if max_rows and r_i > max_rows:
                            buf.write("... (row cap reached)\n")
                            break
                        vals = []
                        for c_i, v in enumerate(row, start=1):
                            if max_cols and c_i > max_cols:
                                break
                            if v is None:
                                vals.append("")
                            else:
                                vals.append(str(v))
                        # trim trailing empties
                        while vals and vals[-1] == "":
                            vals.pop()
                        if vals:
                            buf.write("\t".join(vals) + "\n")
                    buf.write("\n")

                    structured.append({"sheet": sname, "rows_capped": max_rows, "cols_capped": max_cols})

                text = buf.getvalue().strip()
            except Exception as e:
                text = f"[xlsx] failed to parse: {e}"

        return {
            "doc_id": doc_id,
            "text": text,
            "pages": [
                {
                    "page_index": 0,
                    "py_text": text,
                    "text_hint": (text[:200] if text else rel_path),
                    "has_raster": 0,
                    "has_vector": 0,
                    "quality": "queued",
                }
            ],
            "structured_data": structured,
            "meta": {"source": "excel-processor", "rel_path": rel_path},
        }

